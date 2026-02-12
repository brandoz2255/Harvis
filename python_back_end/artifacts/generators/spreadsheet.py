"""
Excel spreadsheet generator using openpyxl
"""

import os
import uuid
import logging
from typing import Dict, Any

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


def generate_spreadsheet(content: Dict[str, Any], output_dir: str) -> str:
    """
    Generate Excel file from manifest content.

    Args:
        content: Spreadsheet content dict with 'sheets' key
        output_dir: Directory to save the file

    Returns:
        Full path to generated file

    Expected content format:
    {
        "sheets": [
            {
                "name": "Sheet1",
                "headers": ["Column A", "Column B"],
                "data": [["Row 1 A", "Row 1 B"], ["Row 2 A", "Row 2 B"]],
                "column_widths": {"A": 20, "B": 30},  # optional
                "freeze_panes": "A2"  # optional
            }
        ],
        "author": "Harvis AI"  # optional
    }
    """
    os.makedirs(output_dir, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    sheets_data = content.get("sheets", [])

    if not sheets_data:
        # Create a default empty sheet
        sheets_data = [{"name": "Sheet1", "headers": [], "data": []}]

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for sheet_data in sheets_data:
        sheet_name = sheet_data.get("name", "Sheet")[:31]  # Excel limit of 31 chars
        ws = wb.create_sheet(title=sheet_name)

        headers = sheet_data.get("headers", [])
        data = sheet_data.get("data", [])
        column_widths = sheet_data.get("column_widths", {})
        freeze_panes = sheet_data.get("freeze_panes")

        # Add headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Add data rows
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=_convert_value(value))
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

        # Auto-fit columns or use specified widths
        for col_idx in range(1, max(len(headers), max((len(row) for row in data), default=0)) + 1):
            col_letter = get_column_letter(col_idx)

            if col_letter in column_widths:
                ws.column_dimensions[col_letter].width = column_widths[col_letter]
            else:
                # Auto-fit based on content
                max_length = 0

                # Check header
                if col_idx <= len(headers):
                    max_length = len(str(headers[col_idx - 1]))

                # Check data
                for row_data in data:
                    if col_idx <= len(row_data):
                        cell_length = len(str(row_data[col_idx - 1] or ""))
                        max_length = max(max_length, cell_length)

                # Set width with some padding
                ws.column_dimensions[col_letter].width = min(max_length + 3, 50)

        # Freeze panes if specified
        if freeze_panes:
            ws.freeze_panes = freeze_panes
        elif headers:
            # Default: freeze header row
            ws.freeze_panes = "A2"

    # Set document properties
    wb.properties.creator = content.get("author", "Harvis AI")

    # Generate unique filename
    filename = f"artifact_{uuid.uuid4().hex[:16]}.xlsx"
    filepath = os.path.join(output_dir, filename)

    wb.save(filepath)
    logger.info(f"Generated spreadsheet: {filepath}")

    return filepath


def _convert_value(value: Any) -> Any:
    """Convert value to appropriate Excel type"""
    if value is None:
        return ""

    if isinstance(value, str):
        # Try to convert numeric strings
        try:
            if "." in value:
                return float(value)
            return int(value)
        except ValueError:
            return value

    return value
